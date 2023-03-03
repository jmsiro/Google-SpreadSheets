# coding: utf-8
"""
Base para desarrollo de modulos externos.
Para obtener el modulo/Funcion que se esta llamando:
     GetParams("module")

Para obtener las variables enviadas desde formulario/comando Rocketbot:
    var = GetParams(variable)
    Las "variable" se define en forms del archivo package.json

Para modificar la variable de Rocketbot:
    SetVar(Variable_Rocketbot, "dato")

Para obtener una variable de Rocketbot:
    var = GetVar(Variable_Rocketbot)

Para obtener la Opcion seleccionada:
    opcion = GetParams("option")


Para instalar librerias se debe ingresar por terminal a la carpeta "libs"

    pip install <package> -t .

"""
import os.path
import sys

base_path = tmp_global_obj["basepath"]
cur_path = base_path + 'modules' + os.sep + \
           'Google-SpreadSheets' + os.sep + 'libs' + os.sep
sys.path.append(cur_path)

from googleapiclient import discovery
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl.utils.cell import get_column_letter

import pickle
import re

"""
    Obtengo el modulo que fueron invocados
"""
module = GetParams("module")

global creds

if module == "GoogleSuite":
    cred = None
    credential_path = GetParams("credentials_path")

    try:
        if not os.path.exists(credential_path):
            raise Exception(
                "El archivo de credenciales no existe en la ruta especificada")

        SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive.file',
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/script.projects',
            'https://www.googleapis.com/auth/script.external_request',
            'https://www.googleapis.com/auth/drive.scripts'
        ]

        if os.path.exists('token_sheet.pickle'):
            with open('token_sheet.pickle', 'rb') as token:
                cred = pickle.load(token)
            # If there are no (valid) credentials available, let the user log in.
        if not cred or not cred.valid:
            if cred and cred.expired and cred.refresh_token:
                cred.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    credential_path, SCOPES)
                cred = flow.run_local_server()
            # Save the credentials for the next run
            with open('token_sheet.pickle', 'wb') as token:
                pickle.dump(cred, token)
        creds = cred
    except Exception as e:
        PrintException()
        raise e

if module == "CreateSpreadSheet":

    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_name = GetParams('ss_name')
    result = GetParams('result')

    service = discovery.build('sheets', 'v4', credentials=creds)

    spreadsheet_body = {
        "properties": {
            "title": ss_name
        }
    }

    request = service.spreadsheets().create(body=spreadsheet_body)
    response = request.execute()
    if result:
        SetVar(result, response["spreadsheetId"])
        
if module == "CreateSheet":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    name = GetParams('name')
    result = GetParams('result')

    service = discovery.build('sheets', 'v4', credentials=creds)

    body = {
        "requests": [
            {
                "addSheet": {
                    "properties": {
                        "title": name,
                    }
                }
            }
        ]
    }

    request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id,
                                                 body=body)
    response = request.execute()

    if result:
        sheetId = response["replies"][0]["addSheet"]["properties"]["sheetId"]
        SetVar(result, sheetId)

if module == "UpdateSheetProperties":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams('sheetName')
    newName = GetParams('newName')
    hidden = GetParams('hidden')
    
    service = discovery.build('sheets', 'v4', credentials=creds)

    data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
    for element in data["sheets"]:
        if element["properties"]["title"] == sheet:
            sheet_id = element["properties"]["sheetId"]
    
    if not 'sheet_id' in locals():
        raise Exception("Sheet could't be found...")
    
    if not newName:
        newName = sheet
    
    if not hidden:
        hidden = False
    else:
        hidden = eval(hidden)
    
    body = {
        "requests": [
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": sheet_id,
                        "title": newName,
                        "hidden": hidden
                        },
                    "fields": "title, hidden",
                    }
                }
            ]
        }

    request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id,
                                                 body=body)
    response = request.execute()

if module == "DeleteSheet":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams('sheetName')
    service = discovery.build('sheets', 'v4', credentials=creds)

    data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
    for element in data["sheets"]:
        if element["properties"]["title"] == sheet:
            sheet_id = element["properties"]["sheetId"]

    if not 'sheet_id' in locals():
        raise Exception("Sheet could't be found...")
    
    body = {
        "requests": [
            {
                "deleteSheet": {
                    "sheetId": sheet_id
                }
            }

        ]
    }

    request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id,
                                                 body=body)
    response = request.execute()

if module == "UpdateRange":

    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    range_ = GetParams('range')
    text = GetParams('text')

    try:
        if not text.startswith("["):
            text = text.replace('"', '\\\"')
            text = "[[ \"{}\" ]]".format(text)
        
        values = eval(text)
        
        service = discovery.build('sheets', 'v4', credentials=creds)

        # Checks existence of the given sheet name and update the range
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                range_ = sheet + "!" + range_ # Sheet1!A1:A10
        
        if not 'range_' in locals():
            raise Exception("Sheet could't be found...")
        
        body = {
            "values": values
        }
        
        request = service.spreadsheets().values().update(spreadsheetId=ss_id, range=range_,
                                                         valueInputOption="USER_ENTERED",
                                                         body=body)
        response = request.execute()
    
    except Exception as e:
        PrintException()
        raise e

def get_column_index(col):
    try:
        abc = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
               'v', 'w', 'x', 'y', 'z']
        around_abc = len(col) - 1
  
        col = col[-1].lower()
        col_index = around_abc * len(abc) + abc.index(col)
        return col_index
    except Exception as e:
        PrintException()
        raise e

if module == "UpdateFormat":
    
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    range_ = GetParams('range')
    merge = GetParams('merge')
    unmerge = GetParams('unmerge')
    resize = GetParams('resize')
    number_format = GetParams('format') # select
    pattern = GetParams('pattern') # string
    
    foreground = GetParams('foreground') # touple
    font_family = GetParams('fontFamily') # string
    font_size = GetParams('fontSize') # int
    
    bold = GetParams('bold') # bool
    italic = GetParams('italic') # bool
    strikethrough = GetParams('strikethrough') # bool
    underline = GetParams('underline') # bool
    
    try:
        
        if ":" in range_:
            range_
        else:
            range_ = range_ + ":" + range_
        
        service = discovery.build('sheets', 'v4', credentials=creds)

        # Checks existence of the given sheet name and update the range
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]

        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
        
        regex = r"([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)"
        range_re = re.findall(regex, range_)
        
        column_start = get_column_index(range_re[0][0])
        column_end = get_column_index(range_re[0][2]) + 1
        
        row_start = int(range_re[0][1]) - 1
        row_end = int(range_re[0][3]) 
        
        body = {
                'requests': [
                    
                ]
            }
        
        if merge:
            if eval(merge) == True:

                merge_ = {
                        "mergeCells": {
                            "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_start,
                            "endRowIndex": row_end,
                            "startColumnIndex": column_start,
                            "endColumnIndex": column_end
                            },
                            "mergeType": "MERGE_ALL"
                        }
                    }
                body['requests'].append(merge_)
        
        if unmerge:
            if eval(unmerge) == True:

                unmerge_ = {
                        "unmergeCells": {
                            "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_start,
                            "endRowIndex": row_end,
                            "startColumnIndex": column_start,
                            "endColumnIndex": column_end
                            }
                        }
                    }
                body['requests'].append(unmerge_)
        
        if resize:
            if eval(resize) == True:   
                columms_ = {
                            'autoResizeDimensions':{
                                'dimensions': {
                                    'sheetId': sheet_id,
                                    'dimension': 'COLUMNS',
                                    'startIndex': column_start,
                                    'endIndex': column_end
                                }  
                            }            
                        }
                body['requests'].append(columms_)

                rows_ = {
                        'autoResizeDimensions':{
                            'dimensions': {
                                'sheetId': sheet_id,
                                'dimension': 'ROWS',
                                'startIndex': row_start,
                                'endIndex': row_end   
                            }  
                        }            
                    }
                body['requests'].append(rows_)
                
        uef = {}
        fields = []
               
        if number_format != '':
            uef['numberFormat'] = {'type': number_format, 'pattern': ''}          
            if pattern != '':
                uef['numberFormat']['pattern'] = pattern     
            fields.append('numberFormat')
            
        if foreground:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            
            c = foreground.split(',')
            # To use RGB format, the API expects a proportion of each over the 255
            uef['textFormat']["foregroundColorStyle"] = {
                "rgbColor": {
                    "red": int(c[0])/255,
                    "green": int(c[1])/255,
                    "blue": int(c[2])/255,
                    "alpha": 1
                }
            }
            
        if font_family != '':
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['fontFamily'] = font_family

        if font_size and eval(font_size) >0:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['fontSize'] = eval(font_size)
        
        if bold:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['bold'] = eval(bold)
            
        if italic:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['italic'] = eval(italic)
        
        if strikethrough:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['strikethrough'] = eval(strikethrough)
        
        if underline:
            if not uef.get('textFormat'):
                uef['textFormat'] = {}
                fields.append('textFormat')
            uef['textFormat']['underline'] = eval(underline)
        
        if fields != []:
        
            fields_ = ','.join(fields)
                
            cell_format= {
                'repeatCell': {
                    'range': {
                        "sheetId": sheet_id,
                        'startRowIndex': row_start,
                        'endRowIndex': row_end,
                        'startColumnIndex': column_start,
                        'endColumnIndex': column_end,
                    },
                    'cell': {
                        'userEnteredFormat': uef
                        },
                    'fields': f'userEnteredFormat({fields_})'
                }
            }
            
            body['requests'].append(cell_format)
        
        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()
            
    except Exception as e:
        PrintException()
        raise e

if module == "ReadCells":

    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    range_ = GetParams('range')
    result = GetParams('result')

    try:
        service = discovery.build('sheets', 'v4', credentials=creds)

        # Checks existence of the given sheet name and update the range
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                range_ = sheet + "!" + range_ # Sheet1!A1:A10
        
        if not 'range_' in locals():
            raise Exception("Sheet could't be found...")
                
        request = service.spreadsheets().values().get(spreadsheetId=ss_id, range=range_)

        response = request.execute()
        try:
            value = response["values"]
        except:
            value = ""

        if result:
            SetVar(result, value)
    except Exception as e:
        PrintException()
        raise e


if module == "copyPaste":
    
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    range_ = GetParams('range')
    
    sheet2 = GetParams("sheetName2")
    range_2 = GetParams('range2')
    
    type_ = GetParams('type')
    transponse = GetParams('transponse')
    cut = GetParams('cut')
       
    try:
        
        if ":" in range_:
            range_
        else:
            range_ = range_ + ":" + range_
        
        if ":" in range_2:
            range_2 
        else:
            range_2 = range_2 + ":" + range_2
        
        service = discovery.build('sheets', 'v4', credentials=creds)

        # Checks existence of the given sheet name and update the range
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
            if element["properties"]["title"] == sheet2:
                sheet_id2 = element["properties"]["sheetId"]

        if not 'sheet_id' in locals():
            raise Exception("Source sheet could't be found...")
        
        if not 'sheet_id2' in locals():
            raise Exception("Target sheet could't be found...")
            
        regex = r"([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)"
        # <----- Data Origin ----->
        range_re = re.findall(regex, range_)
        
        column_start = get_column_index(range_re[0][0])
        column_end = get_column_index(range_re[0][2]) + 1
        
        row_start = int(range_re[0][1]) - 1
        row_end = int(range_re[0][3]) 
        
        # <----- Data Destination ----->
        range_re2 = re.findall(regex, range_2)
        
        column_start2 = get_column_index(range_re2[0][0])
        column_end2 = get_column_index(range_re2[0][2]) + 1
        
        row_start2 = int(range_re2[0][1]) - 1
        row_end2 = int(range_re2[0][3]) 
        
        orientation = "NORMAL"
        if transponse:
            if eval(transponse) == True:
                orientation = 'TRANSPOSE'
        
        body = {
                'requests': [
                    ]
                }
        
        if not cut or eval(cut) == False:
            body['requests'] = {
                        "copyPaste": {
                            "source": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_start,
                            "endRowIndex": row_end,
                            "startColumnIndex": column_start,
                            "endColumnIndex": column_end,
                            },
                            "destination": {
                                "sheetId": sheet_id2,
                                "startRowIndex": row_start2,
                                "endRowIndex": row_end2,
                                "startColumnIndex": column_start2,
                                "endColumnIndex": column_end2,
                                },
                            "pasteType": type_,
                            "pasteOrientation": orientation,
                            }
                        }
        else:
            body['requests'] = {
                    "cutPaste": {
                        "source": {
                        "sheetId": sheet_id,
                        "startRowIndex": row_start,
                        "endRowIndex": row_end,
                        "startColumnIndex": column_start,
                        "endColumnIndex": column_end,
                        },
                        "destination": {
                            "sheetId": sheet_id2,
                            "rowIndex": row_start2,
                            "columnIndex": column_start2,
                            },
                        "pasteType": type_,
                        }
                    }
        
        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()
    
    except Exception as e:
        PrintException()
        raise e
        

if module == "GetSheets":

    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    result = GetParams('result')
    
    service = discovery.build('sheets', 'v4', credentials=creds)
        
    request = service.spreadsheets().get(spreadsheetId=ss_id)
    response = request.execute()

    sheets = []
    for element in response["sheets"]:
        sheets.append(element["properties"]["title"])
    if result:
        SetVar(result, sheets)
        
if module == "CountCells":

    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    try:
        ss_id = GetParams('ss_id')
        sheet = GetParams('sheetName')
        result = GetParams('result') # Here is saved the number of rows the command was originaly made for that.
        columns = GetParams('columns')
        
        range_ = "A1:ZZZ999999"

        service = discovery.build('sheets', 'v4', credentials=creds)
        
        # Checks existence of the given sheet name and update the range
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                range_ = sheet + "!" + range_ # Sheet1!A1:A10
                
        if not 'range_' in locals():
            raise Exception("Sheet could't be found...")
        
        request = service.spreadsheets().values().get(spreadsheetId=ss_id, range=range_)
        response = request.execute()

        length = len(response["values"])
        
        width_aux = max([len(row) for row in response["values"]])
        width = [width_aux, get_column_letter(width_aux)] # get_column_letter indexes begin with 1 (not 0)

        if result:
            SetVar(result, length)

        if columns:
            SetVar(columns, width)
        
    except Exception as e:
        PrintException()
        raise e

if module == "DeleteColumn":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    col = GetParams('column').lower()
    blank = GetParams('blank')

    try:
        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()

        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
        
        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
        
        sep = col.find(":")
        if sep == -1:
            col_index_1 = get_column_index(col)
            col_index_2 = col_index_1 + 1
        else:            
            cols = col.split(":")
            col_index_1 = get_column_index(cols[0])
            col_index_2 = get_column_index(cols[1]) + 1
        
        if blank is not None:
            blank = eval(blank)

        if blank:
            shiftDimension = "ROWS"
        else:
            shiftDimension = "COLUMNS"

        body = {
            "requests": [{
                "deleteRange": {
                    "range": {
                        "sheetId": sheet_id,
                        "startColumnIndex": col_index_1,
                        "endColumnIndex": col_index_2
                    },
                    "shiftDimension": shiftDimension
                }
            }]
        }

        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()

    except Exception as e:
        PrintException()
        raise e

if module == "DeleteRow":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    row = GetParams('row')
    blank = GetParams('blank')

    try:

        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()

        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]

        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
 
        sep = row.find(":")
        if sep == -1:
            row_index_1 = int(row) - 1
            row_index_2 = int(row)
        else:            
            rows = row.split(":")
            row_index_1 = int(rows[0])-1
            row_index_2 = int(rows[1])
        
        if blank is not None:
            blank = eval(blank)

        if blank:
            shiftDimension = "COLUMNS"
        else:
            shiftDimension = "ROWS"

        body = {
            "requests": [{
                "deleteRange": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": row_index_1,
                        "endRowIndex": row_index_2
                    },
                    "shiftDimension": shiftDimension
                }
            }]
        }

        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()

    except Exception as e:
        PrintException()
        raise e

if module == "AddColumn":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    col = GetParams('column').lower()
    q = int(GetParams("q"))
    blank = GetParams('blank')

    try:
        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()

        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]

        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")

        col_index = get_column_index(col)
        
        if blank is not None:
            blank = eval(blank)

        if blank == False or col_index == 0:
            inheritance = "false"
        else:
            inheritance = "true"

        body = {
            "requests": [{
                "insertDimension": {
                    "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_index,
                    "endIndex": col_index + q
                },
                "inheritFromBefore": inheritance
                }
            }]
        }

        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()

    except Exception as e:
        PrintException()
        raise e

if module == "AddRow":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")

    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    row = int(GetParams('row'))
    q = int(GetParams("q"))
    blank = GetParams('blank')
    
    try:

        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()

        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
        
        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
        
        if blank is not None:
            blank = eval(blank)
            
        row_index = row - 1

        if blank == False or row_index == 0:
            inheritance = "false"
        else:
            inheritance = "true"

        body = {
            "requests": [{
                "insertDimension": {
                    "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": row_index,
                    "endIndex": row_index + q
                },
                "inheritFromBefore": inheritance
                }
            }]
        }

        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()

    except Exception as e:
        PrintException()
        raise e

def get_existing_basic_filters(ss_id, service, startRow=0, endRow=1000) -> dict:
    params = {'spreadsheetId': ss_id,
              'fields': 'sheets(properties(sheetId,title),basicFilter)'}
    response = service.spreadsheets().get(**params).execute()

    data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
    values_range = []
    for sheet in response['sheets']:
        if 'basicFilter' in sheet:
            values_range = list(sheet.values())[1]['range']
    # {'startRowIndex': startRow, 'endRowIndex': endRow, 'startColumnIndex': 4, 'endColumnIndex': 5}
    col = chr(values_range['startColumnIndex'] + 65)
    col = col + str(values_range['startRowIndex']+1) + ":" + col + str(values_range['endRowIndex']+1)
    return col

def apply_filters(ss_id, filters, service):
    try:
        # All requests are validated before any are applied, so bundling the set and clear filter
        # operations in the same request would fail: only 1 basic filter can exist at a time.

        def clear_filters(ss_id, known_filters, service):
            requests = []
            for sheetId, filter in known_filters.items():
                requests.append({'clearBasicFilter': {'sheetId': sheetId}})
            if not requests:
                return
            params = {'spreadsheetId': ss_id,
                      'body': {'requests': requests}}
            service.spreadsheets().batchUpdate(**params).execute()

        def removekey(d, key):
            r = dict(d)
            del r[key]
            return r

        clear_filters(ss_id, filters, service)

        requests = []
        for sheetId, filter in filters.items():
            # By removing the starting and ending indices from the 'range' property,
            # we ensure the basicFilter will apply to the entire sheet bounds. If one knows the
            # desired values for startColumnIndex, startRowIndex, endRowIndex, endColumnIndex,
            # then they can be used to create a range-specific basic filter.
            # The 'range' property is a `GridRange`:
            if 'filterSpecs' not in filter:
                filter['filterSpecs'] = [{
                    'filterCriteria': {
                        'hiddenValues': []
                    }
                }]
            requests.append({'setBasicFilter': {'filter': filter}})
        if not requests:
            return
        params = {'spreadsheetId': ss_id,
                  'body': {'requests': requests}}

        service.spreadsheets().batchUpdate(**params).execute()
    except Exception as e:
        PrintException()
        raise e

def create_filter_structure(ranges, values, sheet_id):
    try:
        new_filter = {
            'range': ranges
        }
        startColumnIndex = ranges['startColumnIndex']
        endColumnIndex = ranges['endColumnIndex']
        new_filter['filterSpecs'] = []
        for index in range(startColumnIndex, endColumnIndex):
            new_filter['filterSpecs'].append({
                'columnIndex': index,
                'filterCriteria': {
                    'hiddenValues': values
                }
            })
        new_filter_with_sheet_id = {
            sheet_id: new_filter
        }
        return new_filter_with_sheet_id
    except Exception as e:
        PrintException()
        raise e

if module == "unfilterData":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")
    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    try:
        service = discovery.build('sheets', 'v4', credentials=creds)
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
        if sheet == None or sheet == "":
                    sheet_id = 0
        else:
            for element in data["sheets"]:
                if element["properties"]["title"] == sheet:
                    sheet_id = element["properties"]["sheetId"]

            if not 'sheet_id' in locals():
                raise Exception("Sheet could't be found...")
            
        requests = []
        requests.append({'clearBasicFilter': {'sheetId': sheet_id}})
        
        params = {'spreadsheetId': ss_id,
                  'body': {'requests': requests}}
        
        service.spreadsheets().batchUpdate(**params).execute()      
        
    except Exception as e:
        PrintException()
        raise e

if module == "filterData":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")
    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    col = GetParams("col").lower()
    valor_filtro = GetParams("valor_filtro")
    try:
        col_index = get_column_index(col)

        service = discovery.build('sheets', 'v4', credentials=creds)
        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
        if sheet == None or sheet == "":
            sheet_id = 0
        else:
            for element in data["sheets"]:
                if element["properties"]["title"] == sheet:
                    sheet_id = element["properties"]["sheetId"]

            if not 'sheet_id' in locals():
                raise Exception("Sheet could't be found...")
        
        range_ = sheet+"!A:"+col
        req = service.spreadsheets().values().get(spreadsheetId=ss_id, range=range_).execute()
        
        # It checks where the table that is going to be filtered starts and ends 
        first_row = 0
        values = req["values"]
        for cell in values:
            if cell == []:
                first_row += 1
            else:
                break
        last_row = len(req['values'])
        
        ranges = {
            "sheetId": sheet_id,
            'startRowIndex': first_row,
            'endRowIndex': last_row,
            'startColumnIndex': col_index,
            'endColumnIndex': col_index + 1,
        }
        
        hidden_values = []
        for row in values:
            # It appends a blank space to the list so the row is recognized as one in the following "for"
            if row == []:
                row.append("")
            for cell in row:
                if valor_filtro != cell:
                    hidden_values.append(cell)
        
        filters = create_filter_structure(ranges, hidden_values, sheet_id)
        apply_filters(ss_id, filters, service)
    except Exception as e:
        PrintException()
        raise e

if module == "filterCells":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")
    ss_id = GetParams('ss_id')
    sheet = GetParams("sheetName")
    res = GetParams("res")
    range = GetParams("range_")
    row_info = GetParams("row_info")
    
    try:
        service = discovery.build('sheets', 'v4', credentials=creds)
        
        data_ = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
        for element in data_["sheets"]:
            if sheet == None or sheet == "":
                    sheet = data_["sheets"][0]["properties"]["title"]
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
                
                if not 'sheet_id' in locals():
                    raise Exception("Sheet could't be found...")
                
                filter_start = element["basicFilter"]["range"]["startRowIndex"]
                
        data = service.spreadsheets().get(spreadsheetId=ss_id, fields="sheets(data(rowMetadata(hiddenByFilter)),properties/sheetId)").execute()

        #column_filter = get_existing_basic_filters(ss_id, service)
        list_hidden_rows = []
        for column in data['sheets']:
            if column['properties']['sheetId'] == sheet_id:
                for index, item in enumerate(column['data'][0]['rowMetadata']):
                    if bool(item):
                        list_hidden_rows.append(index)
       
        # It makes sure that always start from the first row of the filter, so the row index vs hidden rows can be done
        range_first_row = range[1]
        
        if range_first_row != filter_start:
            tmp = list(range)
            tmp[1] = filter_start + 1
            
            range = "".join(str(x) for x in tmp)
            
            
        range_ = sheet + "!" + range
        request = service.spreadsheets().values().get(spreadsheetId=ss_id, range=range_)
        response = request.execute()
        value = response["values"]
        
        final_cells = []
        final_cells_row = {}
        for row_index, item in enumerate(value):
            row_index = (row_index + filter_start)
            if row_info and eval(row_info) == True:
                if row_index not in list_hidden_rows:
                    final_cells_row[row_index+1] = item   
                SetVar(res, final_cells_row)
            else:
                if row_index not in list_hidden_rows:
                    final_cells.append(item)                      
                SetVar(res, final_cells)
    except Exception as e:
        PrintException()
        raise e
    
if module == "CopySheet":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")
    try:
        ss_id = GetParams('ss_id')
        sheet = GetParams("sheetName")
        ss_id_2 = GetParams('ss_id_2')
        result = GetParams('res')

        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
        
        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
        
        body =  {
            'destination_spreadsheet_id': ss_id_2
            }
        
        request = service.spreadsheets().sheets().copyTo(spreadsheetId=ss_id, sheetId=sheet_id, body=body)
        response = request.execute()

        SetVar(res, True)
    except Exception as e:
        SetVar(res, False)
        PrintException()
        raise e
    
if module == "TextToColumns":
    if not creds:
        raise Exception(
            "There's no credentials, nor valid token. Please, generate your credentials.")
    try:
        ss_id = GetParams('ss_id')
        sheet = GetParams("sheetName")
        separator = GetParams('separator')
        result = GetParams('res')

        service = discovery.build('sheets', 'v4', credentials=creds)

        data = service.spreadsheets().get(spreadsheetId=ss_id).execute()
        
        for element in data["sheets"]:
            if element["properties"]["title"] == sheet:
                sheet_id = element["properties"]["sheetId"]
        
        if not 'sheet_id' in locals():
            raise Exception("Sheet could't be found...")
        
        body =  {
            'requests':[
                {
                    'textToColumns':  {
                        'source': {
                            'sheetId': sheet_id,
                            'startColumnIndex': 0,
                            'endColumnIndex': 1
                        },
                        'delimiterType': separator
                    }
                }
            ]
        }
        
        request = service.spreadsheets().batchUpdate(spreadsheetId=ss_id, body=body)
        response = request.execute()

        SetVar(res, True)
    except Exception as e:
        SetVar(res, False)
        PrintException()
        raise e